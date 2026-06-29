"""修論発表の座席割り当て最適化 (PuLP/CBC)"""

import csv
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pulp import (
    PULP_CBC_CMD,
    LpBinary,
    LpMaximize,
    LpMinimize,
    LpProblem,
    LpVariable,
    lpSum,
    value,
)

# ── データ読み込み ──────────────────────────────────────────

students = []
committee = {}          # sid -> set(3教員) ※最適化用
supervisor = {}         # sid -> 主査
sub_supervisors = {}    # sid -> [副査1, 副査2]
students_of = defaultdict(list)

with open("student.tsv") as f:
    reader = csv.reader(f, delimiter="\t")
    for row in reader:
        if len(row) < 4:
            continue
        sid, fac_a, fac_b, fac_c = row[0], row[1], row[2], row[3]
        students.append(sid)
        committee[sid] = {fac_a, fac_b, fac_c}
        supervisor[sid] = fac_a
        sub_supervisors[sid] = [fac_b, fac_c]
        for e in [fac_a, fac_b, fac_c]:
            students_of[e].append(sid)

faculties = sorted(students_of.keys())
S = students
E = faculties
V = [0, 1, 2]

print(f"学生数: {len(S)}, 教員数: {len(E)}")
max_load = max(len(students_of[e]) for e in E)
busiest = [e for e in E if len(students_of[e]) == max_load][0]
print(f"最多担当: {busiest} ({max_load}名)")


def solve(R_max):
    R = list(range(R_max))
    print(f"\n{'='*60}")
    print(f"R_max = {R_max} で求解開始")
    print(f"{'='*60}")

    solver = PULP_CBC_CMD(msg=True, timeLimit=120)

    # ── フェーズ1: フル行最大化 ─────────────────────────────

    m1 = LpProblem("phase1_full_rows", LpMaximize)

    x = LpVariable.dicts("x", [(s, r, v) for s in S for r in R for v in V], cat=LpBinary)
    full = LpVariable.dicts("full", R, cat=LpBinary)

    # 目的: フル行最大化
    m1 += lpSum(full[r] for r in R)

    # 制約1: 各学生はちょうど1回配置
    for s in S:
        m1 += lpSum(x[s, r, v] for r in R for v in V) == 1, f"assign_{s}"

    # 制約2: 各セルは1人以下
    for r in R:
        for v in V:
            m1 += lpSum(x[s, r, v] for s in S) <= 1, f"cell_{r}_{v}"

    # 制約3: 教員かぶり禁止
    for e in E:
        for r in R:
            m1 += (
                lpSum(x[s, r, v] for s in students_of[e] for v in V) <= 1,
                f"nodup_{e}_{r}",
            )

    # 制約4: full の定義
    for r in R:
        m1 += lpSum(x[s, r, v] for s in S for v in V) >= 3 * full[r], f"full_{r}"

    # 対称性除去: フル行を前詰め
    for r in R[:-1]:
        m1 += full[r] >= full[r + 1], f"sym_row_{r}"

    print("\n--- フェーズ1: フル行最大化 ---")
    m1.solve(solver)

    if m1.status != 1:
        print(f"フェーズ1 実行不能 (status={m1.status})")
        return None

    T_star = int(round(value(m1.objective)))
    print(f"フェーズ1 最適値: フル行数 T* = {T_star}")

    # ── フェーズ2: 教員の会場移動回数を最小化 ────────────────
    #
    # venue[e,r,v]: 教員eがセッションrで会場vにいる仮想位置。
    #   出席回 → 実際の会場に固定される (venue_link制約)
    #   欠席回 → ソルバーが自由に選び、隣接セッションと揃えて移動を回避する
    # venue が変わった回数 = 会場移動回数 → これを最小化。

    m2 = LpProblem("phase2_transitions", LpMinimize)

    x2 = LpVariable.dicts("x2", [(s, r, v) for s in S for r in R for v in V], cat=LpBinary)
    full2 = LpVariable.dicts("full2", R, cat=LpBinary)
    venue = LpVariable.dicts("venue",
        [(e, r, v) for e in E for r in R for v in V], cat=LpBinary)
    tr = LpVariable.dicts("tr",
        [(e, r) for e in E for r in R[1:]], lowBound=0, upBound=1)
    max_moves = LpVariable("max_moves", lowBound=0)

    # 目的: 合計移動回数を最小化しつつ、同じ合計なら最大移動教員を均等化
    # (合計の係数100 > max_movesの上限16 なので、合計が確実に優先される)
    m2 += 100 * lpSum(tr[e, r] for e in E for r in R[1:]) + max_moves

    # 制約1: 各学生はちょうど1回配置
    for s in S:
        m2 += lpSum(x2[s, r, v] for r in R for v in V) == 1, f"assign_{s}"

    # 制約2: 各セルは1人以下
    for r in R:
        for v in V:
            m2 += lpSum(x2[s, r, v] for s in S) <= 1, f"cell_{r}_{v}"

    # 制約3: 教員かぶり禁止
    for e in E:
        for r in R:
            m2 += (
                lpSum(x2[s, r, v] for s in students_of[e] for v in V) <= 1,
                f"nodup_{e}_{r}",
            )

    # 制約4: full の定義
    for r in R:
        m2 += lpSum(x2[s, r, v] for s in S for v in V) >= 3 * full2[r], f"full_{r}"

    # フェーズ1の最適値を固定
    m2 += lpSum(full2[r] for r in R) == T_star, "fix_full_rows"

    # 対称性除去: フル行を前詰め
    for r in R[:-1]:
        m2 += full2[r] >= full2[r + 1], f"sym_row_{r}"

    # 各教員は各セッションでちょうど1会場にいる（出席・欠席問わず）
    for e in E:
        for r in R:
            m2 += lpSum(venue[e, r, v] for v in V) == 1, f"vsum_{e}_{r}"

    # 出席時リンク: 教員の学生がいる会場に教員もいる
    for e in E:
        soe = students_of[e]
        for r in R:
            for v in V:
                m2 += venue[e, r, v] >= lpSum(x2[s, r, v] for s in soe), f"vl_{e}_{r}_{v}"

    # 移動検出: venue が変わったら tr=1
    for e in E:
        for r in R[1:]:
            for v in V:
                m2 += tr[e, r] >= venue[e, r - 1, v] - venue[e, r, v], f"tr_{e}_{r}_{v}"

    # 均等化: 各教員の移動回数 ≤ max_moves
    for e in E:
        m2 += max_moves >= lpSum(tr[e, r] for r in R[1:]), f"maxm_{e}"

    # 列の対称性除去: 最多担当教員の全学生を会場A(列0)に固定
    for s in students_of[busiest]:
        m2 += lpSum(x2[s, r, v] for r in R for v in [1, 2]) == 0, f"fix_{s}_col0"
    # さらに venue も全セッション会場Aに固定（欠席回含む）
    for r in R:
        m2 += venue[busiest, r, 0] == 1, f"fv_{busiest}_{r}"

    print("\n--- フェーズ2: 教員の会場移動回数最小化 ---")
    solver2 = PULP_CBC_CMD(msg=True, timeLimit=180)
    m2.solve(solver2)

    if m2.status != 1:
        print(f"フェーズ2 未解決 (status={m2.status})")
        if m2.sol_status != 1:
            return None

    total_trans = int(round(value(m2.objective)))
    print(f"フェーズ2 最適値: 会場移動回数 = {total_trans}")

    # ── 解の抽出 ────────────────────────────────────────────

    assignment = {}
    for s in S:
        for r in R:
            for v in V:
                if value(x2[s, r, v]) is not None and value(x2[s, r, v]) > 0.5:
                    assignment[s] = (r, v)

    return assignment, T_star, total_trans


# ── メインループ: R_max を増やしながらリトライ ──────────────

result = None
for R_max in range(20, 30, 2):
    result = solve(R_max)
    if result is not None:
        break

if result is None:
    print("解が見つかりませんでした")
    raise SystemExit(1)

assignment, T_star, total_trans = result

# ── グリッド構築 ────────────────────────────────────────────

VENUE_NAMES = {0: "A", 1: "B", 2: "C"}

grid = defaultdict(dict)
for s, (r, v) in assignment.items():
    grid[r][v] = s

# 行番号を連番に詰め直す（0-indexed の内部行 → 1-indexed の「第N回」）
used_rows_internal = sorted(grid.keys())
row_map = {old: i + 1 for i, old in enumerate(used_rows_internal)}

print(f"\n{'='*70}")
print("  座席配置結果")
print(f"{'='*70}")

# ── グリッド表（標準出力） ──────────────────────────────────

header = f"{'':>6s} | {'会場A':^14s} | {'会場B':^14s} | {'会場C':^14s}"
print(header)
print("-" * len(header))

for old_r in used_rows_internal:
    session = row_map[old_r]
    cells = []
    for v in V:
        cells.append(grid[old_r].get(v, ""))
    line = f"第{session:2d}回 |"
    for c in cells:
        line += f" {c:^14s} |"
    print(line)

# ── 学生一覧表（標準出力） ──────────────────────────────────

print(f"\n{'='*70}")
print("  学生別配置一覧")
print(f"{'='*70}")

list_header = f"{'学生':^12s} | {'主査':^12s} | {'副査1':^12s} | {'副査2':^12s} | {'回':>4s} | {'会場':>4s}"
print(list_header)
print("-" * len(list_header))

sorted_students = sorted(assignment.keys(), key=lambda s: (assignment[s][0], assignment[s][1]))
for s in sorted_students:
    r, v = assignment[s]
    session = row_map[r]
    venue = VENUE_NAMES[v]
    print(
        f"{s:^12s} | {supervisor[s]:^12s} | {sub_supervisors[s][0]:^12s} | "
        f"{sub_supervisors[s][1]:^12s} | {session:>4d} | {venue:>4s}"
    )

# ── サマリ ──────────────────────────────────────────────────

num_sessions = len(used_rows_internal)
row_counts = {r: len(grid[r]) for r in used_rows_internal}
full_count = sum(1 for c in row_counts.values() if c == 3)
two_count = sum(1 for c in row_counts.values() if c == 2)
one_count = sum(1 for c in row_counts.values() if c == 1)

print(f"\n{'='*70}")
print("  サマリ")
print(f"{'='*70}")
print(f"  発表回数:     {num_sessions}回")
print(f"  3会場フル:    {full_count}回")
print(f"  2会場のみ:    {two_count}回")
print(f"  1会場のみ:    {one_count}回")
print(f"  合計所要時間: {num_sessions * 10}分 ({num_sessions}回 × 10分)")

# 教員別の移動回数を計算
faculty_schedule = {}
faculty_transitions = {}

for e in E:
    schedule = []
    for old_r in used_rows_internal:
        for v in V:
            sid = grid[old_r].get(v)
            if sid and e in committee[sid]:
                schedule.append((row_map[old_r], VENUE_NAMES[v]))
                break
    faculty_schedule[e] = schedule
    trans_count = sum(
        1 for i in range(1, len(schedule)) if schedule[i][1] != schedule[i - 1][1]
    )
    faculty_transitions[e] = trans_count

computed_total = sum(faculty_transitions.values())
max_faculty_moves = max(faculty_transitions.values())
print(f"  会場移動回数: {computed_total}回 (全教員合計, 最大{max_faculty_moves}回/人)")

movers = {e: t for e, t in faculty_transitions.items() if t > 0}
print(f"\n  教員別 会場移動詳細 (移動あり: {len(movers)}名 / {len(E)}名):")
print(f"  {'教員':^12s}  {'担当数':>5s}  {'移動':>4s}  会場推移")
print(f"  {'-'*60}")
for e in sorted(E, key=lambda e: (-faculty_transitions[e], e)):
    schedule = faculty_schedule[e]
    if not schedule:
        continue
    venues = [s[1] for s in schedule]
    venue_str = "→".join(venues)
    t = faculty_transitions[e]
    mark = " *" if t > 0 else ""
    print(f"  {e:^12s}  {len(students_of[e]):>5d}  {t:>4d}  {venue_str}{mark}")

# ── 検証 ────────────────────────────────────────────────────

assert len(assignment) == len(S), f"配置学生数 {len(assignment)} != {len(S)}"

for old_r in used_rows_internal:
    row_students = [s for s, (rr, _) in assignment.items() if rr == old_r]
    row_faculties = []
    for s in row_students:
        row_faculties.extend(committee[s])
    assert len(row_faculties) == len(set(row_faculties)), (
        f"第{row_map[old_r]}回で教員が重複: {[e for e in row_faculties if row_faculties.count(e) > 1]}"
    )

print(f"\n  検証OK: 全{len(S)}名配置済み、教員かぶりなし")

# ── TSV保存 ────────────────────────────────────────────────

with open("result.tsv", "w", newline="") as f:
    writer = csv.writer(f, delimiter="\t")
    writer.writerow(["学生", "主査", "副査1", "副査2", "第N回", "会場"])
    for s in sorted_students:
        r, v = assignment[s]
        writer.writerow([s, supervisor[s], sub_supervisors[s][0], sub_supervisors[s][1],
                         row_map[r], VENUE_NAMES[v]])

print("  result.tsv に保存しました")

# ── Excel保存 ──────────────────────────────────────────────

wb = Workbook()

# --- シート1: 学生別配置一覧 ---
ws1 = wb.active
ws1.title = "学生別配置"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
CELL_FONT = Font(size=11)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
VENUE_FILLS = {
    "A": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "B": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "C": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}

headers1 = ["学生", "主査", "副査1", "副査2", "第N回", "会場"]
for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for i, s in enumerate(sorted_students, 2):
    r, v = assignment[s]
    session = row_map[r]
    venue = VENUE_NAMES[v]
    row_data = [s, supervisor[s], sub_supervisors[s][0], sub_supervisors[s][1], session, venue]
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.font = CELL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = VENUE_FILLS.get(venue, PatternFill())

ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 14
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 8
ws1.column_dimensions["F"].width = 8

# --- シート2: タイムテーブル（グリッド表示） ---
ws2 = wb.create_sheet("タイムテーブル")

ws2.cell(row=1, column=1, value="").border = THIN_BORDER
for v in V:
    cell = ws2.cell(row=1, column=v + 2, value=f"会場{VENUE_NAMES[v]}")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for i, old_r in enumerate(used_rows_internal, 2):
    session = row_map[old_r]
    cell = ws2.cell(row=i, column=1, value=f"第{session}回")
    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    for v in V:
        sid = grid[old_r].get(v, "")
        cell = ws2.cell(row=i, column=v + 2, value=sid)
        cell.font = CELL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if sid:
            cell.fill = VENUE_FILLS[VENUE_NAMES[v]]

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14

# --- シート3: タイムテーブル詳細（教員名つき） ---
ws3 = wb.create_sheet("タイムテーブル詳細")

for v in V:
    col_start = v * 4 + 2
    cell = ws3.cell(row=1, column=col_start, value=f"会場{VENUE_NAMES[v]}")
    ws3.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 3)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

cell = ws3.cell(row=1, column=1, value="")
cell.border = THIN_BORDER

sub_headers = ["学生", "主査", "副査1", "副査2"]
for v in V:
    for j, sh in enumerate(sub_headers):
        col = v * 4 + 2 + j
        cell = ws3.cell(row=2, column=col, value=sh)
        cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        cell.font = Font(bold=True, size=10)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

cell = ws3.cell(row=2, column=1, value="")
cell.border = THIN_BORDER

for i, old_r in enumerate(used_rows_internal):
    excel_row = i + 3
    session = row_map[old_r]
    cell = ws3.cell(row=excel_row, column=1, value=f"第{session}回")
    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    cell.font = Font(bold=True, size=10)
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    for v in V:
        col_start = v * 4 + 2
        sid = grid[old_r].get(v, "")
        venue_fill = VENUE_FILLS[VENUE_NAMES[v]] if sid else PatternFill()
        if sid:
            vals = [sid, supervisor[sid], sub_supervisors[sid][0], sub_supervisors[sid][1]]
        else:
            vals = ["", "", "", ""]
        for j, val in enumerate(vals):
            cell = ws3.cell(row=excel_row, column=col_start + j, value=val)
            cell.font = Font(size=10)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if sid:
                cell.fill = venue_fill

ws3.column_dimensions["A"].width = 10
for v in V:
    for j in range(4):
        col_letter = chr(ord("B") + v * 4 + j)
        ws3.column_dimensions[col_letter].width = 13

# --- シート4: 教員スケジュール ---
ws4 = wb.create_sheet("教員スケジュール")

sorted_faculty = sorted(E, key=lambda e: (-faculty_transitions[e], e))

cell = ws4.cell(row=1, column=1, value="")
cell.border = THIN_BORDER
for j, e in enumerate(sorted_faculty, 2):
    cell = ws4.cell(row=1, column=j, value=e)
    cell.fill = HEADER_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=9)
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for i, old_r in enumerate(used_rows_internal, 2):
    session = row_map[old_r]
    cell = ws4.cell(row=i, column=1, value=f"第{session}回")
    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    cell.font = Font(bold=True, size=10)
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    for j, e in enumerate(sorted_faculty, 2):
        venue = ""
        for v in V:
            sid = grid[old_r].get(v)
            if sid and e in committee[sid]:
                venue = VENUE_NAMES[v]
                break
        cell = ws4.cell(row=i, column=j, value=venue)
        cell.font = Font(size=10)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if venue:
            cell.fill = VENUE_FILLS[venue]

# 移動回数サマリ行
summary_row = len(used_rows_internal) + 2
cell = ws4.cell(row=summary_row, column=1, value="移動回数")
cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
cell.font = Font(bold=True, size=10)
cell.alignment = CENTER
cell.border = THIN_BORDER
for j, e in enumerate(sorted_faculty, 2):
    cell = ws4.cell(row=summary_row, column=j, value=faculty_transitions[e])
    cell.font = Font(bold=True, size=10)
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if faculty_transitions[e] > 0:
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

ws4.column_dimensions["A"].width = 10
for j in range(len(sorted_faculty)):
    col_letter = chr(ord("B") + j) if j < 24 else None
    if col_letter:
        ws4.column_dimensions[col_letter].width = 12

wb.save("result.xlsx")
print("  result.xlsx に保存しました")
